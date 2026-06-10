"""Format-specific extractors. Each returns a list of ContentBlock.

Supported on-the-fly:
  .xlsx / .xls   -> one block per meaningful row
  .pdf           -> one block per page
  .txt / .md     -> one block per heading/paragraph group
  .mp4 / .mov / .mkv / .avi / .wav / .m4a -> transcript segments (Vosk, offline)

Adding a format = add one function and register it in `EXTRACTORS`.
Heavy deps (openpyxl, pypdf, vosk) are imported lazily so a missing optional
dependency only breaks its own format, not the whole team.
"""
from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Callable, Optional

from .ckm import ContentBlock

# Callback signature used by all extractors to report each block as it is
# produced (so callers can stream progress instead of waiting for the full list).
OnBlock = Optional[Callable[[ContentBlock], None]]


def _emit(blocks: list[ContentBlock], block: ContentBlock, on_block: OnBlock) -> None:
    blocks.append(block)
    if on_block is not None:
        try:
            on_block(block)
        except Exception:
            pass

# Where extracted visuals (pdf figures, video frames) are written. The
# ExtractionAgent points this at <outputs>/assets before running.
ASSETS_DIR = Path(__file__).resolve().parent / "outputs" / "assets"

# Cap on frames grabbed from a single media file (evenly sampled) so a long
# video doesn't produce thousands of images.
_MEDIA_MAX_FRAMES = int(os.getenv("MEDIA_MAX_FRAMES", "60"))


def _slug(*parts: object) -> str:
    raw = "__".join(str(p) for p in parts)
    return re.sub(r"[^a-zA-Z0-9_]+", "-", raw).strip("-").lower()


# ---------- xlsx ----------

def extract_xlsx(path: Path, on_block: OnBlock = None) -> list[ContentBlock]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    blocks: list[ContentBlock] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        for i, row in enumerate(rows[1:], start=1):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if not cells:
                continue
            pairs = {
                header[j]: str(v).strip()
                for j, v in enumerate(row)
                if j < len(header) and header[j] and v is not None and str(v).strip()
            }
            _emit(blocks, ContentBlock(
                id=_slug(path.stem, sheet.title, i),
                source=path.name,
                modality="table_row",
                title=f"{sheet.title} row {i}",
                text=" | ".join(f"{k}: {v}" for k, v in pairs.items()) or " | ".join(cells),
                metadata={"sheet": sheet.title, "fields": pairs},
            ), on_block)
    wb.close()
    return blocks


# ---------- pdf ----------

def extract_pdf(path: Path, on_block: OnBlock = None) -> list[ContentBlock]:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    blocks: list[ContentBlock] = []
    for page_no, page in enumerate(reader.pages, start=1):
        text = (page.extract_text() or "").strip()
        image_ref = _save_pdf_page_image(path, page, page_no)
        if not text and not image_ref:
            continue
        _emit(blocks, ContentBlock(
            id=_slug(path.stem, "p", page_no),
            source=path.name,
            modality="text",
            title=f"{path.stem} — page {page_no}",
            text=text,
            image_ref=image_ref,
            metadata={"page": page_no},
        ), on_block)
    return blocks


def _save_pdf_page_image(pdf: Path, page, page_no: int) -> str | None:
    """Save the largest embedded image on a PDF page; return its path or None."""
    try:
        images = list(getattr(page, "images", []))
    except Exception:
        return None
    if not images:
        return None
    best = max(images, key=lambda im: len(getattr(im, "data", b"") or b""))
    if not getattr(best, "data", None):
        return None
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    ext = Path(getattr(best, "name", "img.png")).suffix or ".png"
    out = ASSETS_DIR / f"{_slug(pdf.stem, 'p', page_no)}{ext}"
    try:
        out.write_bytes(best.data)
    except Exception:
        return None
    return str(out)


# ---------- text / markdown ----------

def extract_text(path: Path, on_block: OnBlock = None) -> list[ContentBlock]:
    content = path.read_text(encoding="utf-8", errors="ignore")
    blocks: list[ContentBlock] = []
    chunks = re.split(r"\n(?=#{1,6}\s)", content)  # split on markdown headings
    for i, chunk in enumerate(chunks):
        chunk = chunk.strip()
        if not chunk:
            continue
        first = chunk.splitlines()[0].lstrip("# ").strip()
        _emit(blocks, ContentBlock(
            id=_slug(path.stem, i),
            source=path.name,
            modality="heading" if chunk.startswith("#") else "text",
            title=first[:80],
            text=chunk,
            metadata={},
        ), on_block)
    return blocks


# ---------- audio / video ----------

def extract_media(path: Path, on_block: OnBlock = None) -> list[ContentBlock]:
    import os
    import tempfile
    import wave
    import json as _json
    import subprocess

    # Reuse a cached transcript if the source is unchanged (transcription is slow).
    cache = path.with_suffix(path.suffix + ".transcript.json")
    if cache.exists() and cache.stat().st_mtime >= path.stat().st_mtime:
        raw = _json.loads(cache.read_text(encoding="utf-8"))
        blocks = [ContentBlock(**b) for b in raw]
        # Backfill visuals if the cache predates frame extraction.
        if not any(b.image_ref for b in blocks):
            _grab_frames(path, blocks)
            cache.write_text(
                _json.dumps([b.model_dump() for b in blocks], ensure_ascii=False),
                encoding="utf-8",
            )
        if on_block is not None:
            for b in blocks:
                try:
                    on_block(b)
                except Exception:
                    pass
        return blocks

    from vosk import KaldiRecognizer, Model, SetLogLevel

    SetLogLevel(-1)
    model_dir = Path(os.getenv(
        "VOSK_MODEL_PATH",
        str(Path(__file__).resolve().parent.parent / "tetris_mvp" / "models" / "vosk-model-small-en-us-0.15"),
    ))
    if not model_dir.exists():
        raise FileNotFoundError(f"Vosk model not found at {model_dir} (set VOSK_MODEL_PATH)")

    with tempfile.TemporaryDirectory() as tmp:
        wav_path = Path(tmp) / "audio.wav"
        if path.suffix.lower() in {".wav"}:
            wav_path = path
        else:
            import imageio_ffmpeg

            ffmpeg = imageio_ffmpeg.get_ffmpeg_exe()
            subprocess.run(
                [ffmpeg, "-y", "-i", str(path), "-ac", "1", "-ar", "16000",
                 "-vn", "-f", "wav", str(wav_path)],
                check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )

        model = Model(str(model_dir))
        blocks: list[ContentBlock] = []
        with wave.open(str(wav_path), "rb") as wf:
            rec = KaldiRecognizer(model, wf.getframerate())
            rec.SetWords(True)

            def _flush(res_json: str) -> None:
                r = _json.loads(res_json)
                text = (r.get("text") or "").strip()
                if not text:
                    return
                words = r.get("result") or []
                start = round(words[0]["start"], 2) if words else 0.0
                _emit(blocks, ContentBlock(
                    id=_slug(path.stem, "t", int(start * 100)),
                    source=path.name,
                    modality="transcript",
                    title=f"{path.stem} @ {start:.0f}s",
                    text=text,
                    timestamp=start,
                ), on_block)

            while True:
                data = wf.readframes(4000)
                if not data:
                    break
                if rec.AcceptWaveform(data):
                    _flush(rec.Result())
            _flush(rec.FinalResult())

    _grab_frames(path, blocks)
    cache.write_text(
        _json.dumps([b.model_dump() for b in blocks], ensure_ascii=False),
        encoding="utf-8",
    )
    return blocks


def _grab_frames(video: Path, blocks: list[ContentBlock]) -> None:
    """Sample up to _MEDIA_MAX_FRAMES transcript blocks and grab a video frame
    at each timestamp, setting block.image_ref. No-op for pure audio."""
    if video.suffix.lower() in {".wav", ".m4a", ".mp3"} or not blocks:
        return
    import subprocess
    import imageio_ffmpeg

    timed = [b for b in blocks if b.timestamp is not None]
    if not timed:
        return
    step = max(1, len(timed) // _MEDIA_MAX_FRAMES)
    sampled = timed[::step]
    ffmpeg = imageio_ffmpeg.get_ffmpeg_exe()
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    for b in sampled:
        # print(b.timestamp)
        out = ASSETS_DIR / f"{b.id}.jpg"
        if not (out.exists() and out.stat().st_size > 0):
            try:
                subprocess.run(
                    [ffmpeg, "-y", "-ss", f"{max(0.0, b.timestamp):.2f}", "-i", str(video),
                     "-frames:v", "1", "-q:v", "3", str(out)],
                    check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                )
            except subprocess.CalledProcessError:
                continue
        if out.exists() and out.stat().st_size > 0:
            b.image_ref = str(out)


EXTRACTORS = {
    ".xlsx": extract_xlsx,
    ".xls": extract_xlsx,
    ".pdf": extract_pdf,
    ".txt": extract_text,
    ".md": extract_text,
    ".mp4": extract_media,
    ".mov": extract_media,
    ".mkv": extract_media,
    ".avi": extract_media,
    ".wav": extract_media,
    ".m4a": extract_media,
}


def extract_file(path: Path, on_block: OnBlock = None) -> list[ContentBlock]:
    fn = EXTRACTORS.get(path.suffix.lower())
    if not fn:
        return []
    return fn(path, on_block)
