"""xlsx / xls source tool: one ContentBlock per meaningful row."""
from __future__ import annotations

from pathlib import Path

from agent_team.core.ckm import ContentBlock

from ._base import OnBlock, _emit, _slug


def extract_xlsx(path: Path, on_block: OnBlock = None) -> list[ContentBlock]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    blocks: list[ContentBlock] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        for i, row in enumerate(rows[1:], start=1):
            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if not cells:
                continue
            pairs = {
                header[j]: str(v).strip()
                for j, v in enumerate(row)
                if j < len(header) and header[j] and v is not None and str(v).strip()
            }
            _emit(blocks, ContentBlock(
                id=_slug(path.stem, sheet.title, i),
                source=path.name,
                modality="table_row",
                title=f"{sheet.title} row {i}",
                text=" | ".join(f"{k}: {v}" for k, v in pairs.items()) or " | ".join(cells),
                metadata={"sheet": sheet.title, "fields": pairs},
            ), on_block)
    wb.close()
    return blocks
